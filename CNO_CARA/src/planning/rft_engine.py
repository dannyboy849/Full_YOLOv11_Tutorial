#!/usr/bin/env python3
"""
remaining flight time (RFT) predictor for VTOL-capable aircraft.

generated automatically by uas-modeling when the customer Excel export was created.
reads calibrated SysID parameters from sheet "SysID_Params" in ONSSI-320-00000043_Processed_Output.xlsx.

usage:
    python ONSSI-320-00000043_RFT_Predictor.py --workbook ONSSI-320-00000043_Processed_Output.xlsx --soc 0.5

dependencies:
    pip install openpyxl numpy
"""

from __future__ import annotations

import math
import numpy        as np
import argparse

from pathlib        import Path
from typing         import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Install openpyxl: pip install openpyxl") from exc


# ---------------------------------------------------------------------------
# embedded defaults from calibration at export time. Workbook rows override
# ---------------------------------------------------------------------------

DEFAULT_PARAMS: dict[str, dict[str, Any]] = {'vehicle': {'profile': 'ONSSI-320', 'mass_kg': 21.5}, 'fw': {'cd0': np.float64(0.020174307695619963), 'k': np.float64(0.029650632649987296), 'eta': 0.75, 'eta_fixed': True, 'fit_eta': False, 'reynolds_ref': 537739.8154403435, 'uses_altitude_density': True, 'altitude_ref_m': 832.1499633789062, 'n_fit_samples': 12566, 'uses_reynolds': True}, 'vtol': {'model': 'phys_v1', 'p_idle_w': 620.1168083611827, 'k_lift_w': 266.36394764349285, 'k_lift_v_w_per_ms': 3.988315848458451e-09, 'k_drag_w_per_ms3': 1.865445103231798e-17, 'k_push_w': 142.102229145682, 'lift_exp': 1.0000000000003375, 'lift_min': 0.29913762759591694, 'lift_span': 0.612214009151707, 'b0': 508.6643825825818, 'b1': 415.18862184955213, 'b2': -0.18068439645571743, 'rmse_w': 106.27927209921842, 'n_fit': 818}, 'battery': {'capacity_ah': 27.0, 'voltage_nominal_v': 44.4, 'avionics_power_w': 50.0, 'wing_area_m2': 0.8, 'rho_kgm3': 1.225}, 'rft_scenario': {'target_ground_speed_mps': 21.0, 'wind_speed_mps': 0.0, 'wind_angle_deg': 0.0, 'vtol_lift_proxy': 0.25, 'vtol_airspeed_mps': 5.0, 'vtol_pusher_proxy': 0.0, 'state_of_charge': 0.5}, 'fit_quality': {'fw_mape_percent': 21.885139270623498, 'fw_train_mape_percent': 20.443602303358457, 'fw_validation_mape_percent': 21.885139270623498, 'fw_n_fit_samples': 12566, 'fw_n_train_samples': 12566, 'fw_n_validation_samples': 3142, 'vtol_mape_percent': 30.50384139681701, 'vtol_train_mape_percent': 12.013338120378345, 'vtol_validation_mape_percent': 30.50384139681701, 'vtol_n_fit_samples': 818}}

DEFAULT_WORKBOOK = Path(__file__).resolve().parent / "ONSSI-320-00000043_Processed_Output.xlsx"
KNOTS_TO_MPS = 1852.0 / 3600.0
MPS_TO_KNOTS = 3600.0 / 1852.0


def knots_to_mps(knots: float) -> float:
    return float(knots) * KNOTS_TO_MPS


def mps_to_knots(mps: float) -> float:
    return float(mps) * MPS_TO_KNOTS


def resolve_rft_speeds(target_ground_speed: float, wind_speed: float, unit: str = "mps") -> tuple[float, float]:
    normalized = (unit or "mps").strip().lower()
    if normalized in {"knot", "knots", "kt", "kts"}:
        return knots_to_mps(target_ground_speed), knots_to_mps(wind_speed)
    return float(target_ground_speed), float(wind_speed)



# ---------------------------------------------------------------------------
# METHODOLOGY
# ---------------------------------------------------------------------------
# Fixed-wing (FW) remaining flight time
# ------------------------------------
# 1. Estimate electrical power P_fw for level flight at a target ground speed V_g,
#    given wind speed V_w and wind direction (direction wind blows toward).
# 2. Solve the wind triangle for true airspeed V_a:
#        V_g^2 = V_a^2 + V_w^2 + 2 V_a V_w cos(theta)
#    and keep physically valid roots (positive airspeed).
# 3. Aerodynamic power from a calibrated drag polar:
#        C_D = C_D0 + k C_L^2
#        D = 0.5 rho V_a^2 S C_D
#        P_elec = (D V_a) / eta + P_avionics
# 4. RFT_fw (min) = (E_remaining_Wh / P_fw_W) * 60
#
# VTOL remaining flight time
# -------------------------
# VTOL hover/transition power is modeled separately from wing-borne cruise.
# The calibrated VTOL map gives electrical power vs lift_proxy, airspeed, and
# optional pusher command. At a representative operating point (sheet
# rft_scenario or CLI flags):
#        RFT_vtol (min) = (E_remaining_Wh / P_vtol_W) * 60
#
# This is a steady operating-point estimate, not a full mission simulation.
# Use FW RFT for cruise planning and VTOL RFT for hover/transition budgeting.
# ---------------------------------------------------------------------------


def load_params_from_workbook(workbook_path: Path) -> dict[str, dict[str, Any]]:
    """
    Load sectioned parameters from the SysID_Params worksheet.

    Expected columns: section | parameter | value | unit | notes

    Returns nested dict {section: {parameter: value}} merged onto DEFAULT_PARAMS.
    """
    merged = {k: dict(v) for k, v in DEFAULT_PARAMS.items()}
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if "SysID_Params" not in wb.sheetnames:
        wb.close()
        return merged
    ws = wb["SysID_Params"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    for row in rows:
        if not row or len(row) < 3:
            continue
        section = str(row[0]).strip() if row[0] is not None else ""
        name = str(row[1]).strip() if row[1] is not None else ""
        value = row[2]
        if not section or not name:
            continue
        merged.setdefault(section, {})
        if isinstance(value, (int, float)) or value is None:
            merged[section][name] = value
        else:
            text = str(value).strip()
            if text.lower() in {"true", "yes"}:
                merged[section][name] = True
            elif text.lower() in {"false", "no"}:
                merged[section][name] = False
            else:
                try:
                    merged[section][name] = float(text)
                except ValueError:
                    merged[section][name] = text
    return merged


def _get(params: dict[str, dict[str, Any]], section: str, name: str, default: Any) -> Any:
    """Fetch a parameter with fallback to embedded defaults."""
    return params.get(section, {}).get(name, DEFAULT_PARAMS.get(section, {}).get(name, default))


def calculate_cl(airspeed_mps: float, mass_kg: float, wing_area_m2: float, rho: float) -> float:
    """
    Lift coefficient for level flight: L = W = m g and L = 0.5 rho V^2 S C_L.
    """
    weight_n = mass_kg * 9.80665
    q = 0.5 * rho * max(airspeed_mps, 0.01) ** 2
    return weight_n / (q * wing_area_m2)


def _skin_friction_coeff(reynolds: float) -> float:
    re = max(float(reynolds), 1e-9)
    cf_laminar = 1.328 / math.sqrt(re)
    cf_turbulent = 0.074 / (re ** 0.2)
    return cf_laminar if re < 5e5 else cf_turbulent


def _cd0_effective(cd0: float, reynolds: float, re_ref: float | None) -> float:
    if re_ref is None or re_ref <= 0:
        return cd0
    cf = _skin_friction_coeff(reynolds)
    cf_ref = _skin_friction_coeff(re_ref)
    scale = max(0.25, min(4.0, cf / cf_ref))
    return cd0 * scale


def predict_fw_power_w(airspeed_mps: float, params: dict[str, dict[str, Any]]) -> float:
    """
    Predict fixed-wing electrical power (watts) at a given true airspeed.

    Uses calibrated C_D0, k, eta, wing area, density, and avionics load.
    """
    wing_area = float(_get(params, "battery", "wing_area_m2", 0.8))
    rho = float(_get(params, "battery", "rho_kgm3", 1.225))
    avionics = float(_get(params, "battery", "avionics_power_w", 50.0))
    cd0 = float(_get(params, "fw", "cd0", 0.02))
    k = float(_get(params, "fw", "k", 0.04))
    eta = float(_get(params, "fw", "eta", 0.75))
    mass_kg = float(_get(params, "vehicle", "mass_kg", 21.5))
    uses_reynolds = bool(_get(params, "fw", "uses_reynolds", False))
    re_ref = _get(params, "fw", "reynolds_ref", None)
    chord_m = float(_get(params, "vehicle", "mean_chord_m", 0.34))
    mu = float(_get(params, "battery", "mu_pa_s", 1.81e-5))

    v = max(float(airspeed_mps), 0.5)
    if uses_reynolds and re_ref is not None:
        re = (rho * v * chord_m) / max(mu, 1e-9)
        cd0 = _cd0_effective(cd0, re, float(re_ref))
    cl = calculate_cl(v, mass_kg, wing_area, rho)
    dynamic_pressure = 0.5 * rho * v * v
    cd = cd0 + k * (cl ** 2)
    drag = dynamic_pressure * wing_area * cd
    return (drag * v) / max(eta, 1e-3) + avionics


def predict_power_for_ground_speed_w(
    target_ground_speed_mps: float,
    wind_speed_mps: float,
    wind_angle_deg: float,
    params: dict[str, dict[str, Any]],
) -> float:
    """
    FW electrical power (watts) to sustain target ground speed in wind.

    Solves the wind triangle and returns the maximum power among valid roots
    (conservative for endurance planning).
    """
    v_w = float(wind_speed_mps)
    theta = math.radians(float(wind_angle_deg))
    v_g_target = float(target_ground_speed_mps)

    a = 1.0
    b = 2.0 * v_w * math.cos(theta)
    c = v_w * v_w - v_g_target * v_g_target
    disc = b * b - 4.0 * a * c
    if disc < 0.0:
        return float("inf")

    sqrt_disc = math.sqrt(disc)
    roots = [(-b + sqrt_disc) / (2.0 * a), (-b - sqrt_disc) / (2.0 * a)]
    tol = max(0.05 * v_g_target, 0.25)
    powers: list[float] = []
    for v_air in roots:
        if not math.isfinite(v_air) or v_air <= 0.0:
            continue
        v_g_chk = math.sqrt(v_air * v_air + v_w * v_w + 2.0 * v_air * v_w * math.cos(theta))
        if abs(v_g_chk - v_g_target) > tol:
            continue
        powers.append(predict_fw_power_w(v_air, params))
    return max(powers) if powers else float("inf")


def _normalized_lift_proxy(lift_proxy: float, lift_min: float, lift_span: float) -> float:
    """Map lift_proxy to 0..1 using VTOL SysID normalization."""
    span = lift_span if lift_span > 1e-9 else 1.0
    return float(np.clip((lift_proxy - lift_min) / span, 0.0, 1.0))


def predict_vtol_power_w(
    lift_proxy: float,
    airspeed_mps: float,
    params: dict[str, dict[str, Any]],
    pusher_proxy: float = 0.0,
) -> float:
    """
    Predict VTOL electrical power (watts) at a representative operating point.

    phys_v1: P = P_idle + k_lift*L^exp + k_lift_v*L*V + k_drag*V^3 + k_push*push
    polynomial fallback: b0 + b1*L + b2*L*V
    """
    section = params.get("vtol", {})
    model = str(section.get("model", DEFAULT_PARAMS.get("vtol", {}).get("model", "phys_v1"))).lower()
    lift_n = _normalized_lift_proxy(
        float(lift_proxy),
        float(section.get("lift_min", 0.0)),
        float(section.get("lift_span", 1.0)),
    )
    v = max(float(airspeed_mps), 0.0)
    push_n = float(np.clip(pusher_proxy, 0.0, 1.0))

    if model == "phys_v1":
        p_idle = float(section.get("p_idle_w", 0.0))
        k_lift = float(section.get("k_lift_w", 0.0))
        k_lift_v = float(section.get("k_lift_v_w_per_ms", 0.0))
        k_drag = float(section.get("k_drag_w_per_ms3", 0.0))
        k_push = float(section.get("k_push_w", 0.0))
        exp = float(section.get("lift_exp", 1.5))
        return p_idle + k_lift * (lift_n ** exp) + k_lift_v * lift_n * v + k_drag * (v ** 3) + k_push * push_n

    b0 = float(section.get("b0", 0.0))
    b1 = float(section.get("b1", 0.0))
    b2 = float(section.get("b2", 0.0))
    return b0 + b1 * float(lift_proxy) + b2 * float(lift_proxy) * max(v, 0.5)


def remaining_energy_wh(
    params: dict[str, dict[str, Any]],
    state_of_charge: float | None,
    integrated_ah: float | None,
) -> float:
    """
    Remaining energy (Wh) from SOC or consumed amp-hours.

    SOC uses nominal pack voltage. Coulomb counting uses linear cell voltage vs SOC.
    """
    capacity_ah = float(_get(params, "battery", "capacity_ah", 27.0))
    v_nom = float(_get(params, "battery", "voltage_nominal_v", 44.4))
    if integrated_ah is not None:
        remaining_ah = max(0.0, capacity_ah - float(integrated_ah))
        soc = remaining_ah / capacity_ah
        cells = 12.0
        v_cell_min = 3.2
        v_cell_max = 4.2
        v_pack = cells * (v_cell_min + (v_cell_max - v_cell_min) * soc)
        v_empty = cells * v_cell_min
        v_avg = 0.5 * (v_pack + v_empty)
        return remaining_ah * v_avg
    soc_val = 0.5 if state_of_charge is None else float(state_of_charge)
    return capacity_ah * v_nom * soc_val


def predict_fw_rft_minutes(
    params: dict[str, dict[str, Any]],
    *,
    state_of_charge: float | None,
    integrated_ah: float | None,
    target_ground_speed_mps: float | None,
    wind_speed_mps: float | None,
    wind_angle_deg: float | None,
) -> tuple[float, float, float]:
    """
    FW remaining flight time (minutes), power (W), and energy (Wh).
    """
    energy_wh = remaining_energy_wh(params, state_of_charge, integrated_ah)
    v_g = float(
        target_ground_speed_mps
        if target_ground_speed_mps is not None
        else _get(params, "rft_scenario", "target_ground_speed_mps", 21.0)
    )
    v_w = float(wind_speed_mps if wind_speed_mps is not None else _get(params, "rft_scenario", "wind_speed_mps", 0.0))
    angle = float(wind_angle_deg if wind_angle_deg is not None else _get(params, "rft_scenario", "wind_angle_deg", 0.0))
    power_w = predict_power_for_ground_speed_w(v_g, v_w, angle, params)
    if not math.isfinite(power_w) or power_w <= 0.0:
        return float("inf"), power_w, energy_wh
    return (energy_wh / power_w) * 60.0, power_w, energy_wh


def predict_vtol_rft_minutes(
    params: dict[str, dict[str, Any]],
    *,
    state_of_charge: float | None,
    integrated_ah: float | None,
    lift_proxy: float | None,
    airspeed_mps: float | None,
    pusher_proxy: float | None,
) -> tuple[float, float, float]:
    """
    VTOL remaining flight time (minutes) at a reference operating point.
    """
    energy_wh = remaining_energy_wh(params, state_of_charge, integrated_ah)
    lp = float(lift_proxy if lift_proxy is not None else _get(params, "rft_scenario", "vtol_lift_proxy", 0.25))
    v = float(airspeed_mps if airspeed_mps is not None else _get(params, "rft_scenario", "vtol_airspeed_mps", 5.0))
    push = float(pusher_proxy if pusher_proxy is not None else _get(params, "rft_scenario", "vtol_pusher_proxy", 0.0))
    power_w = predict_vtol_power_w(lp, v, params, pusher_proxy=push)
    if not math.isfinite(power_w) or power_w <= 0.0:
        return float("inf"), power_w, energy_wh
    return (energy_wh / power_w) * 60.0, power_w, energy_wh


def main() -> int:
    """Load workbook parameters and print FW + VTOL RFT estimates."""
    parser = argparse.ArgumentParser(
        description="Predict FW and VTOL remaining flight time from SysID workbook."
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="Customer export .xlsx containing sheet SysID_Params",
    )
    parser.add_argument("--soc", type=float, default=None, help="State of charge 0..1")
    parser.add_argument("--integrated-ah", type=float, default=None, help="Amp-hours consumed since full")
    parser.add_argument("--target-ground-speed", type=float, default=None, help="FW target ground speed")
    parser.add_argument("--wind-speed", type=float, default=None, help="Wind speed")
    parser.add_argument(
        "--speed-units",
        choices=["mps", "knots"],
        default="mps",
        help="Units for --target-ground-speed and --wind-speed (default: m/s).",
    )
    parser.add_argument("--wind-angle", type=float, default=None, help="Wind angle (deg)")
    parser.add_argument("--vtol-lift-proxy", type=float, default=None, help="VTOL lift_proxy reference")
    parser.add_argument("--vtol-airspeed", type=float, default=None, help="VTOL reference airspeed (m/s)")
    parser.add_argument("--vtol-pusher-proxy", type=float, default=None, help="VTOL pusher command 0..1")
    args = parser.parse_args()

    params = load_params_from_workbook(args.workbook.resolve())
    target_default = _get(params, "rft_scenario", "target_ground_speed_mps", 21.0)
    wind_default = _get(params, "rft_scenario", "wind_speed_mps", 0.0)
    target_input = float(args.target_ground_speed if args.target_ground_speed is not None else target_default)
    wind_input = float(args.wind_speed if args.wind_speed is not None else wind_default)
    target_mps, wind_mps = resolve_rft_speeds(target_input, wind_input, unit=args.speed_units)
    fw_rft, fw_p, e_wh = predict_fw_rft_minutes(
        params,
        state_of_charge=args.soc,
        integrated_ah=args.integrated_ah,
        target_ground_speed_mps=target_mps,
        wind_speed_mps=wind_mps,
        wind_angle_deg=args.wind_angle,
    )
    vtol_rft, vtol_p, _ = predict_vtol_rft_minutes(
        params,
        state_of_charge=args.soc,
        integrated_ah=args.integrated_ah,
        lift_proxy=args.vtol_lift_proxy,
        airspeed_mps=args.vtol_airspeed,
        pusher_proxy=args.vtol_pusher_proxy,
    )

    print(f"Workbook: {args.workbook}")
    print(f"Remaining energy: {e_wh:.1f} Wh")
    print()
    print("Fixed-wing cruise (wind triangle + drag polar)")
    print(f"  Ground speed: {target_mps:.2f} m/s ({mps_to_knots(target_mps):.1f} kt)")
    print(f"  Wind speed: {wind_mps:.2f} m/s ({mps_to_knots(wind_mps):.1f} kt)")
    print(f"  Electrical power: {fw_p:.1f} W")
    if math.isfinite(fw_rft):
        print(f"  Remaining flight time: {fw_rft:.1f} min")
    else:
        print("  Remaining flight time: undefined")
    print()
    print("VTOL reference operating point")
    print(f"  Electrical power: {vtol_p:.1f} W")
    if math.isfinite(vtol_rft):
        print(f"  Remaining flight time: {vtol_rft:.1f} min")
    else:
        print("  Remaining flight time: undefined")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
